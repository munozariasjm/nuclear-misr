import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import joblib
import os
import numpy as np
import sys

path_2_this_file = os.path.abspath(os.path.dirname(__file__))
home = os.path.abspath(f"{path_2_this_file}/../")

sys.path.append(home)
from nuclearpy_models.models import sr_be, sr_rc, dz_be, mnp_rc


PATH_ASSETS = f"{home}/gui/assets"
PATH_TH_DATA = f"{home}/Data/Theory/MasterNuclei.xlsx"
th2DATA = pd.read_excel(PATH_TH_DATA)
PATH_EXP_DATA = f"{home}/Data/Experimental/AME2020.csv"
PATH_RC_DATA = f"{home}/Data/Experimental/rc.csv"
PATH_FRDM_DATA = f"{home}/Data/Theory/FRDM2012.csv"
exp_data = pd.read_csv(PATH_EXP_DATA)
rc_data = pd.read_csv(PATH_RC_DATA)
frdm_data = pd.read_csv(PATH_FRDM_DATA)
PATH_MODEL = f"{home}/Data/model/be_model.pkl"
model = joblib.load(PATH_MODEL)


def get_sheets_name(filename):
    workbook = load_workbook(filename)
    return workbook.sheetnames


th_datastes = get_sheets_name(PATH_TH_DATA)


def get_datasets(used_dfs):
    th_datastes = [th_dataste for th_dataste in used_dfs if th_dataste != "AME2020"]
    thdfs = {
        th_dataste: pd.read_excel(PATH_TH_DATA, sheet_name=th_dataste)
        .query("Z >= 12")
        .query("Z<50")
        for th_dataste in th_datastes
    }
    return thdfs


def inference_be(Z, N, index, thdfs, data):
    be_misr_pred = sr_be(Z, N, index)
    binding_energy_preds = {
        # "Z": Z,
        # "N": N,
        "Exp": data["BE"].values[0],
        "MISR": be_misr_pred[0],
        "unc_MISR": be_misr_pred[1],
        "DZ": dz_be(Z, N),
    }
    # prediction for the binding energy with saved model
    ard_p = model.predict(
        [[binding_energy_preds["MISR"], binding_energy_preds["DZ"], Z, N]],
        return_std=True,
    )
    print(ard_p)
    binding_energy_preds["ARD"] = ard_p[0][0]
    binding_energy_preds["unc_ARD"] = ard_p[1][0]
    print(binding_energy_preds)
    try:
        binding_energy_preds["FRDM"] = (
            frdm_data.query("Z == @Z").query("N == @N")["BE"].values[0]
        )
    except:
        binding_energy_preds["FRDM"] = None
    try:
        binding_energy_preds["Exp"] = (
            data.query("Z == @Z").query("N == @N")["BE"].values[0]
        )
        binding_energy_preds["unc_exp"] = (
            data.query("Z == @Z").query("N == @N")["uBE"].values[0]
        )
    except:
        binding_energy_preds["Exp"] = None
        binding_energy_preds["unc_exp"] = None
    for th_dataste in thdfs:
        thdf = thdfs[th_dataste]
        thdf = thdf.query("Z == @Z").query("N == @N")
        if thdf.empty:
            continue
        else:
            binding_energy_preds[th_dataste] = thdf["BE"].values[0]

    return pd.Series(binding_energy_preds, name="Predictions")


def inference_rc(Z, N, index, thdfs, data):
    rc_misr_pred = sr_rc(Z, N, index)
    charge_radii_preds = {
        # "Z": Z,
        # "N": N,
        "MISR": rc_misr_pred[0],
        "unc_MISR": rc_misr_pred[1],
        "MNP": mnp_rc(Z, N),
    }
    try:
        charge_radii_preds["Exp"] = (
            data.query("Z == @Z").query("N == @N")["Rav"].values[0]
        )
        charge_radii_preds["unc_exp"] = (
            data.query("Z == @Z").query("N == @N")["delta_Rav"].values[0]
        )
    except:
        charge_radii_preds["Exp"] = None
        charge_radii_preds["unc_exp"] = None
    for th_dataste in thdfs:
        thdf = thdfs[th_dataste]
        thdf = thdf.query("Z == @Z").query("N == @N")
        if thdf.empty:
            continue
        else:
            charge_radii_preds[th_dataste] = thdf["ChRad"].values[0]
    return pd.Series(charge_radii_preds, name="Predictions")


def plot_values(series, title):
    fig, ax = plt.subplots()
    colors = plt.cm.jet(np.linspace(0, 1, len(series) + 5))
    # Fill the values for experimental data if is not None
    if "ARD" in series:
        ax.errorbar(
            [-1],
            series["ARD"],
            yerr=series["unc_ARD"],
            fmt="o",
            label="ARD",
            color=colors[0],
        )
    if "Exp" in series and series["Exp"] is not None:
        try:
            ax.errorbar(
                [-2],
                series["Exp"],
                yerr=series["unc_exp"],
                # fmt=".r",
                label="Exp",
                color="black",
            )
            ax.fill_between(
                list(range(-3, len(series) + 3)),
                [series["Exp"] - series["unc_exp"]]
                * (len(list(range(-3, len(series) + 3)))),
                [series["Exp"] + series["unc_exp"]]
                * (len(list(range(-3, len(series) + 3)))),
                color="black",
                alpha=0.5,
            )
        except Exception as e:
            print(e)
    # Fill the values for theoretical data
    for i, th_dataste in enumerate(series.index):
        if th_dataste not in ["Exp", "unc_exp", "unc_MISR", "ARD", "unc_ARD", "MISR"]:
            if series[th_dataste] is not None:
                ax.scatter(
                    [i], series[th_dataste], label=th_dataste, color=colors[i + 1]
                )
    # error bar for MISR
    ax.errorbar(
        [0],
        series["MISR"],
        yerr=series["unc_MISR"],
        fmt="o",
        label="MISR",
        color=colors[-1],
    )
    ax.set_ylabel(title)
    # remove xticks
    non_unc = [i for i in series.index if ("unc" not in i)]
    values = [series[i] for i in non_unc if isinstance(series[i], float)]
    ax.set(
        xlim=(-3, len(series) + 1),
        ylim=(
            np.min(values) - 1 * np.std(values),
            np.max(values) + 1 * np.std(values),
        ),
    )
    ax.set_xticks([])
    ax.legend()
    st.pyplot(fig)


def main():
    st.title("Discovering Nuclear Models from Symbolic Machine Learning")

    st.write(
        "This is a web application to inference on the MISR and ARD models for the binding energy and charge radii of atomic nuclei."
    )
    st.write("Read the paper [here](https://arxiv.org/abs/2404.11477)")
    st.sidebar.title("Models")
    cols_input = st.columns(3)
    with cols_input[0]:
        Z = st.number_input("Enter the number of protons (Z)", 12, 50, 18)
    with cols_input[1]:
        N = st.number_input("Enter the number of neutrons (N)", 12, 50, 20)
    with cols_input[2]:
        # select an index with slider
        index = st.slider("Select MISR's index", 0, 10, 10)
    if Z < 12 or Z > 50:
        st.warning("Z must be between 12 and 50. Models are trained for this range.")
        return

    st.sidebar.header("About")
    st.sidebar.image(f"{PATH_ASSETS}/ema.png", use_column_width=True)
    st.sidebar.image(f"{PATH_ASSETS}/mit.png", use_column_width=True)

    # Now display the models
    st.sidebar.header("Select Models")
    possible_theory_models = [
        # "DD-ME2",
        # The line `# "SKMS",` is a commented-out line of code in the list
        # `possible_theory_models`. This means that the model "SKMS" is not
        # currently selected for use in the application. It is part of the list of
        # possible theory models that can be selected by the user, but it is
        # currently not being included in the models that will be used for the
        # calculations.
        # "SKMS",
        "NL3S",
        "UNEDF1",
    ]
    used_dfs = st.sidebar.multiselect(
        "Select the datasets to use",
        th_datastes,
        possible_theory_models[:2],
    )
    with cols_input[1]:
        run_ = st.button("Run Models")
    if run_:
        with st.spinner("Running models..."):
            thdfs = get_datasets(used_dfs)
            st.sidebar.write("The following datasets are used:")
            st.sidebar.write(used_dfs)

            be_pred = inference_be(Z, N, index, thdfs, exp_data)
            rc_pred = inference_rc(Z, N, index, thdfs, rc_data)
            st.divider()
            cols = st.columns(2)
            with cols[0]:
                st.write("Binding Energy [MeV]")
                st.write(be_pred)
                # plot_values(be_pred, "Binding Energy")

            with cols[1]:
                st.write("Charge Radii [fm]")
                st.write(rc_pred)
                # plot_values(rc_pred, "Charge Radii")
            ccols = st.columns(2)
            st.divider()
            with ccols[0]:
                try:
                    plot_values(be_pred, "Binding Energy [MeV]")
                except Exception as e:
                    print(e)
            with ccols[1]:
                try:
                    plot_values(rc_pred, "Charge Radii [fm]")
                except Exception as e:
                    print(e)
            # st.divider()
            st.sidebar.write(
                """The data for the DFT models was taken from the MasterNuclei dataset
                compilated by the [BMex project](https://bmex.dev/)."""
            )


if __name__ == "__main__":
    main()
